from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction 
from .models import Job, Project, Personil, AsetMesin, JobDate, CustomUser, LeaveEvent, Karyawan
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
import requests
import json
# Tambahkan 'Count' dan 'Case' untuk kalkulasi
from django.db.models import Q, Count, Case, When, IntegerField 
from django.urls import reverse
import datetime 
import calendar
from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .utils import format_tanggal_id, get_month_name_id, calculate_daily_jobs_summary, calculate_project_jobs_summary, generate_pdf_filename 
from django.templatetags.static import static
from django.conf import settings
from django.core.paginator import Paginator

from .forms import (
    PersonilForm, 
    ProjectForm, 
    JobDateStatusForm, 
    JobForm, 
    AttachmentFormSet,
    LeaveEventForm,
)

# ==============================================================================
# VIEW HALAMAN UTAMA (DASHBOARD) - (ROMBAK BESAR)
# ==============================================================================
@login_required(login_url='core:login')
def dashboard_view(request):
    user = request.user
    
    # === 1. LOGIKA FILTER (BULAN & TAHUN) ===
    now = datetime.datetime.now()
    try:
        current_year_param = request.GET.get('year', '')
        current_month_param = request.GET.get('month', '')
        
        # Jika kosong atau '0', gunakan nilai default
        # Default: tampilkan "Semua" (0) ATAU jika user pilih hanya 1, default ke tahun sekarang
        if current_year_param and current_year_param != '0':
            current_year = int(current_year_param)
        else:
            current_year = 0  # Artinya "Semua tahun"
        
        if current_month_param and current_month_param != '0':
            current_month = int(current_month_param)
            # Jika user pilih bulan tapi tidak pilih tahun, default ke tahun sekarang
            if current_year == 0:
                current_year = now.year
        else:
            current_month = 0  # Artinya "Semua bulan"
    except ValueError:
        current_year = 0
        current_month = 0
    
    # Flag untuk menentukan apakah filter aktif atau "Semua"
    # Jika KEDUANYA 0, berarti "Semua". Jika salah satu ada value, filter aktif.
    filter_all_dates = (current_month == 0 and current_year == 0)
    
    year_list = range(now.year - 2, now.year + 3) 
    month_list = [
        {"id": 1, "name": "Januari"}, {"id": 2, "name": "Februari"},
        {"id": 3, "name": "Maret"}, {"id": 4, "name": "April"},
        {"id": 5, "name": "Mei"}, {"id": 6, "name": "Juni"},
        {"id": 7, "name": "Juli"}, {"id": 8, "name": "Agustus"},
        {"id": 9, "name": "September"}, {"id": 10, "name": "Oktober"},
        {"id": 11, "name": "November"}, {"id": 12, "name": "Desember"}
    ]
    
    # === 2. LOGIKA FILTER PIC (KONSEP BARU ANDA) ===
    subordinate_ids = user.get_all_subordinates()
    subordinates_list = CustomUser.objects.filter(id__in=subordinate_ids).order_by('username')
    
    selected_pic_id = request.GET.get('pic', '')
    
    if selected_pic_id == 'my_jobs':
        team_query = Q(pic=user) | Q(assigned_to=user)
    elif selected_pic_id:
        try:
            selected_user = get_object_or_404(CustomUser, id=int(selected_pic_id))
            selected_user_sub_ids = selected_user.get_all_subordinates()
            team_ids = selected_user_sub_ids + [selected_user.id]
            team_query = Q(pic_id__in=team_ids) | Q(assigned_to_id__in=team_ids)
        except (ValueError, CustomUser.DoesNotExist):
            team_query = (Q(pic=user) | Q(assigned_to=user)) | (Q(pic_id__in=subordinate_ids) | Q(assigned_to_id__in=subordinate_ids)) # Fallback
    else:
        team_query = (Q(pic=user) | Q(assigned_to=user)) | (Q(pic_id__in=subordinate_ids) | Q(assigned_to_id__in=subordinate_ids))

    # === 3. LOGIKA FILTER ASET (BARU) ===
    selected_line_id = request.GET.get('line', '')
    selected_mesin_id = request.GET.get('mesin', '')
    selected_sub_mesin_id = request.GET.get('sub_mesin', '')
    
    # === 4. LOGIKA DATA TABEL (QUERY OPTIMASI BARU) ===
    # Filter Job berdasarkan TIM, TAPI JANGAN filter tanggal dulu
    all_jobs_team_base = Job.objects.filter(team_query).distinct()
    
    # Apply asset filters
    if selected_sub_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset_id=selected_sub_mesin_id)
    elif selected_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent_id=selected_mesin_id)
    elif selected_line_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent__parent_id=selected_line_id)
    
    # Sekarang filter job yang punya tanggal di bulan/tahun terpilih
    # Jika filter_all_dates = True, ambil semua job tanpa filter bulan/tahun
    if filter_all_dates:
        all_jobs_team = all_jobs_team_base.select_related(
            'pic', 
            'assigned_to',
            'project',
            'aset__parent__parent' 
        ).prefetch_related(
            'personil_ditugaskan', 
            'tanggal_pelaksanaan',
            'attachments'
        ).distinct()
    else:
        # Build filter dynamically based on what user selected
        date_filter = Q()
        if current_month != 0:
            date_filter &= Q(tanggal_pelaksanaan__tanggal__month=current_month)
        if current_year != 0:
            date_filter &= Q(tanggal_pelaksanaan__tanggal__year=current_year)
        
        all_jobs_team = all_jobs_team_base.filter(date_filter).select_related(
            'pic', 
            'assigned_to',
            'project',
            'aset__parent__parent' 
        ).prefetch_related(
            'personil_ditugaskan', 
            'tanggal_pelaksanaan',
            'attachments'
        ).distinct()

    # === 3a. LOGIKA SORTING BARU ===
    sort_by = request.GET.get('sort', 'updated_at')  # Default ke 'updated_at'
    sort_order = request.GET.get('order', 'desc')  # Default descending
    
    # Mapping kolom sorting ke field database
    sort_mapping = {
        'nama_pekerjaan': 'nama_pekerjaan',
        'pic': 'pic__username',
        'line': 'aset__parent__parent__nama',
        'mesin': 'aset__parent__nama',
        'sub_mesin': 'aset__nama',
        'fokus': 'fokus',
        'prioritas': 'prioritas',
        'updated_at': 'updated_at',
    }
    
    # Dapatkan field sorting yang valid
    sort_field = sort_mapping.get(sort_by, 'updated_at')
    
    # Tambahkan prefix - untuk descending
    if sort_order == 'desc':
        sort_field = f'-{sort_field}'
    
    # Apply sorting (kecuali untuk 'progress' yang perlu special handling)
    if sort_by == 'progress':
        # Untuk progress, kita akan sort di Python setelah query
        # karena ini adalah calculated field
        daily_job_data = list(all_jobs_team.filter(tipe_job='Daily'))
        project_job_data = list(all_jobs_team.filter(tipe_job='Project'))
        
        # Sort berdasarkan get_progress_percent()
        reverse_sort = (sort_order == 'desc')
        daily_job_data.sort(key=lambda x: x.get_progress_percent(), reverse=reverse_sort)
        project_job_data.sort(key=lambda x: x.get_progress_percent(), reverse=reverse_sort)
    else:
        # Sorting normal via database
        daily_job_data = all_jobs_team.filter(tipe_job='Daily').order_by(sort_field)
        project_job_data = all_jobs_team.filter(tipe_job='Project').order_by(sort_field)
        
    modal_form = JobDateStatusForm()

    # === 5. LOGIKA DROPDOWN DATA ASET ===
    # Get all Lines (level=0)
    line_list = AsetMesin.objects.filter(level=0).order_by('nama')
    
    # Get Mesin based on selected Line (level=1)
    mesin_list = AsetMesin.objects.none()
    if selected_line_id:
        mesin_list = AsetMesin.objects.filter(parent_id=selected_line_id, level=1).order_by('nama')
    
    # Get Sub Mesin based on selected Mesin (level=2)
    sub_mesin_list = AsetMesin.objects.none()
    if selected_mesin_id:
        sub_mesin_list = AsetMesin.objects.filter(parent_id=selected_mesin_id, level=2).order_by('nama')
    
    # === 6. LOGIKA BARU: DASBOR PROGRES MESIN (KONSEP 2 ANDA) ===
    
    # Ambil semua Aset (Mesin, level=1) yang relevan dengan tim Anda
    # dan punya job di bulan/tahun terpilih (atau semua jika filter_all_dates)
    if filter_all_dates:
        relevant_aset_ids = all_jobs_team_base.filter(
            aset__level=2, # Ambil Sub-Mesin
        ).values_list('aset__parent_id', flat=True).distinct() # Ambil ID Mesin (parent-nya)
    else:
        aset_date_filter = Q()
        if current_month != 0:
            aset_date_filter &= Q(tanggal_pelaksanaan__tanggal__month=current_month)
        if current_year != 0:
            aset_date_filter &= Q(tanggal_pelaksanaan__tanggal__year=current_year)
        
        relevant_aset_ids = all_jobs_team_base.filter(
            aset__level=2 # Ambil Sub-Mesin
        ).filter(aset_date_filter).values_list('aset__parent_id', flat=True).distinct() # Ambil ID Mesin (parent-nya)
    
    relevant_mesin = AsetMesin.objects.filter(id__in=relevant_aset_ids)

    # Hitung progres untuk setiap mesin
    progress_data = []
    for mesin in relevant_mesin:
        # Ambil semua JobDate yang terkait dengan mesin ini & filter
        if filter_all_dates:
            dates_in_mesin = JobDate.objects.filter(
                job__in=all_jobs_team_base,      # Hanya job dari tim
                job__aset__parent=mesin,         # Hanya job di bawah mesin ini
            )
        else:
            date_filter_mesin = Q()
            if current_month != 0:
                date_filter_mesin &= Q(tanggal__month=current_month)
            if current_year != 0:
                date_filter_mesin &= Q(tanggal__year=current_year)
            
            dates_in_mesin = JobDate.objects.filter(
                job__in=all_jobs_team_base,      # Hanya job dari tim
                job__aset__parent=mesin,         # Hanya job di bawah mesin ini
            ).filter(date_filter_mesin)
        
        total_dates = dates_in_mesin.count()
        if total_dates > 0:
            done_dates = dates_in_mesin.filter(status='Done').count()
            progress = int((done_dates / total_dates) * 100)
            progress_data.append({
                'nama': mesin.nama,
                'progress': progress
            })
    
    # Urutkan berdasarkan progress (opsional, tapi bagus)
    progress_data = sorted(progress_data, key=lambda x: x['progress'], reverse=True)
    
    # === 7. PAGINATION LOGIC (BARU) ===
    # Get page size from GET parameter, default 20
    try:
        page_size = int(request.GET.get('page_size', 20))
        if page_size not in [10, 20, 30, 40, 50, 100]:
            page_size = 20
    except (ValueError, TypeError):
        page_size = 20
    
    # Get current page numbers
    try:
        daily_page = int(request.GET.get('daily_page', 1))
        project_page = int(request.GET.get('project_page', 1))
    except (ValueError, TypeError):
        daily_page = 1
        project_page = 1
    
    # Paginate daily jobs
    daily_paginator = Paginator(daily_job_data, page_size)
    daily_page_obj = daily_paginator.get_page(daily_page)
    
    # Paginate project jobs
    project_paginator = Paginator(project_job_data, page_size)
    project_page_obj = project_paginator.get_page(project_page)

    # ==========================================================

    context = {
        'user': user,
        'daily_job_data': daily_page_obj.object_list,
        'daily_page_obj': daily_page_obj,
        'daily_paginator': daily_paginator,
        'daily_total_count': daily_paginator.count,
        
        'project_job_data': project_page_obj.object_list,
        'project_page_obj': project_page_obj,
        'project_paginator': project_paginator,
        'project_total_count': project_paginator.count,
        
        'page_size': page_size,
        'page_size_options': [10, 20, 30, 40, 50, 100],
        
        # Filter Context
        'current_month': current_month,
        'current_year': current_year,
        'month_list': month_list,
        'year_list': year_list,
        'modal_form': modal_form, 
        
        'subordinates_list': subordinates_list,
        'selected_pic_id': selected_pic_id,
        
        # Asset Filter Context (BARU)
        'line_list': line_list,
        'mesin_list': mesin_list,
        'sub_mesin_list': sub_mesin_list,
        'selected_line_id': selected_line_id,
        'selected_mesin_id': selected_mesin_id,
        'selected_sub_mesin_id': selected_sub_mesin_id,
        
        # Sorting Context (BARU)
        'sort_by': sort_by,
        'sort_order': sort_order,
        
        'filter_params': request.GET.urlencode(),
        
        'progress_data': progress_data, # <-- KIRIM DATA PROGRES BARU
    }
    return render(request, 'dashboard.html', context)


# ==============================================================================
# VIEW DETAIL PROJECT (QUERY OPTIMASI BARU)
# ==============================================================================
@login_required(login_url='core:login')
def project_detail_view(request, project_id):
    user = request.user
    project = get_object_or_404(Project, id=project_id)
    
    # === PERMISSION CHECK: Filter jobs berdasarkan user ===
    # User bisa lihat job jika:
    # 1. User adalah PIC (pembuat)
    # 2. User adalah assigned_to
    # 3. User adalah supervisor dari PIC
    # 4. User adalah supervisor dari assigned_to
    subordinate_ids = user.get_all_subordinates()
    
    jobs_in_project = project.jobs.all().filter(
        Q(pic=user) |  # Pembuat
        Q(assigned_to=user) |  # Assigned to
        Q(pic_id__in=subordinate_ids) |  # Supervisor dari PIC
        Q(assigned_to_id__in=subordinate_ids)  # Supervisor dari assigned_to
    ).select_related(
        'pic', 
        'assigned_to',
        'aset__parent__parent' 
    ).prefetch_related(
        'personil_ditugaskan', 
        'tanggal_pelaksanaan',
        'attachments'
    ).distinct()

    modal_form = JobDateStatusForm()

    context = {
        'project': project,
        'job_data': jobs_in_project, 
        'modal_form': modal_form,
        'filter_params': request.GET.urlencode(),
    }
    return render(request, 'detail_project.html', context)


# ==============================================================================
# VIEW MANAJEMEN PERSONIL (TIM SAYA)
# ==============================================================================
@login_required(login_url='core:login')
def manajemen_personil(request):
    user = request.user
    personil_to_edit = None
    edit_id = request.GET.get('edit')
    if edit_id:
        personil_to_edit = get_object_or_404(Personil, id=edit_id, penanggung_jawab=user)
    if request.method == 'POST':
        form = PersonilForm(request.POST, instance=personil_to_edit) if personil_to_edit else PersonilForm(request.POST)
        if form.is_valid():
            personil = form.save(commit=False) 
            personil.penanggung_jawab = user 
            personil.save() 
            messages.success(request, f"Sukses menyimpan data {personil.nama_lengkap}!")
            return redirect('core:manajemen_personil')
        else:
            messages.error(request, "Gagal menyimpan, cek error di form.")
    else:
        form = PersonilForm(instance=personil_to_edit) if personil_to_edit else PersonilForm()
    personil_list = Personil.objects.filter(penanggung_jawab=user).order_by('nama_lengkap')
    context = {'form': form, 'personil_list': personil_list}
    return render(request, 'manajemen_personil.html', context)

# ==============================================================================
# VIEW HAPUS PERSONIL
# ==============================================================================
@login_required(login_url='core:login')
def delete_personil(request, personil_id):
    personil = get_object_or_404(Personil, id=personil_id, penanggung_jawab=request.user)
    try:
        nama = personil.nama_lengkap
        personil.delete()
        messages.success(request, f"Sukses menghapus data {nama}.")
    except Exception as e:
        messages.error(request, f"Gagal menghapus data: {e}")
    return redirect('core:manajemen_personil')

# ==============================================================================
# VIEW MANAJEMEN PROJECT
# ==============================================================================
@login_required(login_url='core:login')
def manajemen_project(request):
    user = request.user
    project_to_edit = None
    edit_id = request.GET.get('edit')
    if edit_id:
        project_to_edit = get_object_or_404(Project, id=edit_id)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project_to_edit) if project_to_edit else ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            if not project_to_edit:
                project.manager_project = user 
            project.save() 
            messages.success(request, f"Sukses menyimpan project: {project.nama_project}!")
            return redirect('core:manajemen_project')
        else:
            messages.error(request, "Gagal menyimpan, cek error di form.")
    else:
        form = ProjectForm(instance=project_to_edit) if project_to_edit else ProjectForm()
    project_list = Project.objects.all().order_by('-created_at')
    context = {'form': form, 'project_list': project_list}
    return render(request, 'manajemen_project.html', context)

# ==============================================================================
# VIEW HAPUS PROJECT
# ==============================================================================
@login_required(login_url='core:login')
def delete_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    try:
        nama = project.nama_project
        project.delete()
        messages.success(request, f"Sukses menghapus project: {nama}.")
    except Exception as e:
        messages.error(request, f"Gagal menghapus: {e}. Mungkin project ini sudah dipakai di Job.")
    return redirect('core:manajemen_project')

# ==============================================================================
# VIEW UPDATE STATUS TANGGAL (MODAL POP-UP)
# ==============================================================================
@login_required(login_url='core:login')
def update_job_date_status(request, job_date_id):
    user = request.user
    subordinate_ids = user.get_all_subordinates()
    
    # Izinkan update jika user adalah:
    # - PIC (pembuat job)
    # - assigned_to (yang ditugaskan)
    # - supervisor dari PIC (PIC ada di bawah user)
    # - supervisor dari assigned_to (assigned_to ada di bawah user)
    allowed_users_query = (
        Q(job__pic=user) |
        Q(job__assigned_to=user) |
        Q(job__pic_id__in=subordinate_ids) |
        Q(job__assigned_to_id__in=subordinate_ids)
    )

    job_date = get_object_or_404(JobDate, allowed_users_query, id=job_date_id)
    
    if request.method == 'POST':
        form = JobDateStatusForm(request.POST, instance=job_date)
        if form.is_valid():
            form.save()
            messages.success(request, f"Sukses update status tanggal {job_date.tanggal}!")
        else:
            messages.error(request, "Gagal update status.")
            
    referer = request.META.get('HTTP_REFERER', reverse('core:dashboard'))
    return HttpResponseRedirect(referer)


# ==============================================================================
# VIEW FORM JOB (TAMBAH / EDIT)
# ==============================================================================
@login_required(login_url='core:login')
def job_form_view(request, job_id=None, project_id=None): 
    user = request.user
    instance = None
    project = None
    
    if job_id:
        job = get_object_or_404(Job, id=job_id)
        # === LOGIKA: Cek permission untuk edit ===
        # Bisa edit HANYA jika:
        # 1. User adalah pembuat (pic) ATAU
        # 2. User adalah assigned_to ATAU
        # 3. User adalah supervisor dari assigned_to
        subordinate_ids = user.get_all_subordinates()
        
        can_edit = (
            job.pic == user or  # Pembuat
            job.assigned_to == user or  # Assigned_to sendiri
            (job.assigned_to is not None and job.assigned_to.id in subordinate_ids) or  # Atasan dari assigned_to
            (job.pic is not None and job.pic.id in subordinate_ids)  # Atasan dari PIC (supervisor of PIC)
        )
        
        if not can_edit:
            messages.error(request, "Anda tidak memiliki akses untuk mengedit job ini.")
            return redirect('core:dashboard')
        
        instance = job 
    
    if project_id:
        project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        form = JobForm(request.POST, instance=instance, user=user, project=project)
        attachment_formset = AttachmentFormSet(request.POST, request.FILES, instance=instance, prefix='attachments')
        
        sub_mesin_id = request.POST.get('sub_mesin')
        
        if form.is_valid() and attachment_formset.is_valid() and sub_mesin_id:
            try:
                with transaction.atomic():
                    # === PENTING: Save original assigned_to value sebelum form process ===
                    original_assigned_to = None
                    if instance:
                        original_assigned_to = instance.assigned_to
                    
                    job = form.save(commit=False)
                    # Hanya set PIC saat create (bukan edit)
                    if not instance:
                        job.pic = user 
                    job.aset_id = sub_mesin_id 
                    
                    # === LOGIKA: Kontrol assigned_to ===
                    # Hanya PIC (pembuat) yang bisa ubah assigned_to
                    # Jika subordinate (assigned_to) yang edit, restore assigned_to value
                    if instance:  # Edit mode
                        if job.pic == user:
                            # PIC bisa ubah assigned_to dari form
                            assigned_to_id = request.POST.get('assigned_to')
                            if assigned_to_id:
                                try:
                                    job.assigned_to_id = int(assigned_to_id)
                                except (ValueError, TypeError):
                                    job.assigned_to = None
                            else:
                                job.assigned_to = None
                        else:
                            # Subordinate (assigned_to) yang edit - RESTORE assigned_to!
                            job.assigned_to = original_assigned_to
                    else:  # Create mode
                        # Saat create, PIC bisa set assigned_to
                        # HANYA update jika form memiliki field assigned_to (yang hanya ada untuk supervisors)
                        if 'assigned_to' in form.fields:
                            assigned_to_id = request.POST.get('assigned_to')
                            if assigned_to_id:
                                try:
                                    job.assigned_to_id = int(assigned_to_id)
                                except (ValueError, TypeError):
                                    job.assigned_to = None
                            else:
                                job.assigned_to = None
                        # Else: jangan ubah assigned_to (tetap None untuk create)
                    
                    if not instance:
                        job.status = 'Open' 
                    
                    if project:
                        job.project = project
                        job.tipe_job = 'Project'
                    
                    job.save() 
                    form.save_m2m()
                    
                    tanggal_str = form.cleaned_data.get('tanggal_pelaksanaan', '')
                    
                    if instance:
                        # Edit mode
                        if tanggal_str:
                            # Jika ada tanggal baru, REPLACE semua dengan yang baru
                            job.tanggal_pelaksanaan.all().delete()  # Hapus tanggal lama
                            tanggal_list = tanggal_str.split(',') 
                            for tgl_str in tanggal_list:
                                if tgl_str: 
                                    tgl_obj = datetime.datetime.strptime(tgl_str.strip(), '%Y-%m-%d').date()
                                    JobDate.objects.create(job=job, tanggal=tgl_obj, status='Open')
                        # Else: jika tanggal kosong saat edit, KEEP existing tanggal (jangan hapus)
                    else:
                        # Create mode
                        if tanggal_str:
                            tanggal_list = tanggal_str.split(',') 
                            for tgl in tanggal_list:
                                if tgl: 
                                    JobDate.objects.create(job=job, tanggal=tgl.strip(), status='Open')
                    
                    attachment_formset.instance = job
                    attachment_formset.save()

                messages.success(request, f"Sukses menyimpan pekerjaan: {job.nama_pekerjaan}")
                
                if project:
                    return redirect('core:project_detail', project_id=project.id)
                return redirect('core:dashboard') 
                
            except Exception as e:
                messages.error(request, f"Terjadi kesalahan saat menyimpan: {e}")
        else:
            if not sub_mesin_id and request.POST.get('line'):
                form.add_error(None, "Data Aset (Line, Mesin, Sub Mesin) wajib diisi lengkap.")
            elif not form.is_valid():
                 messages.error(request, "Gagal menyimpan. Harap periksa error di form.")
            else:
                 messages.error(request, "Gagal menyimpan. Harap periksa lampiran.")

    else:
        form = JobForm(instance=instance, user=user, project=project)
        attachment_formset = AttachmentFormSet(instance=instance, prefix='attachments')
        
        if instance:
            dates_str = ",".join([
                str(d.tanggal) for d in instance.tanggal_pelaksanaan.all()
            ])
            form.fields['tanggal_pelaksanaan'].initial = dates_str

    context = {
        'form': form,
        'attachment_formset': attachment_formset,
        'project': project,
    }
    return render(request, 'job_form.html', context)

# ==============================================================================
# VIEW HAPUS JOB
# ==============================================================================
@login_required(login_url='core:login')
def job_delete(request, job_id):
    user = request.user
    job = get_object_or_404(Job, id=job_id)
    
    # === LOGIKA: Cek permission untuk delete ===
    # Bisa delete HANYA jika:
    # 1. User adalah pembuat (pic) ATAU
    # 2. User adalah assigned_to ATAU
    # 3. User adalah supervisor dari assigned_to (bukan dari pic!)
    subordinate_ids = user.get_all_subordinates()
    
    can_delete = (
        job.pic == user or  # Pembuat
        job.assigned_to == user or  # Assigned_to sendiri
        (job.assigned_to is not None and job.assigned_to.id in subordinate_ids) or  # Atasan dari assigned_to
        (job.pic is not None and job.pic.id in subordinate_ids)  # Atasan dari PIC (supervisor of PIC)
    )
    
    if not can_delete:
        messages.error(request, "Anda tidak memiliki akses untuk menghapus job ini.")
        referer = request.META.get('HTTP_REFERER', reverse('core:dashboard'))
        return HttpResponseRedirect(referer)
    
    try:
        nama = job.nama_pekerjaan
        job.delete()
        messages.success(request, f"Sukses menghapus pekerjaan: {nama}")
    except Exception as e:
        messages.error(request, f"Gagal menghapus: {e}")
        
    referer = request.META.get('HTTP_REFERER', reverse('core:dashboard'))
    return HttpResponseRedirect(referer)

# ==============================================================================
# VIEW AJAX UNTUK DROPDOWN ASET
# ==============================================================================
def load_children(request):
    parent_id = request.GET.get('parent_id')
    try:
        children = AsetMesin.objects.filter(parent_id=parent_id).order_by('nama')
        data = [{"id": child.id, "nama": child.nama} for child in children]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# ==============================================================================
# VIEW AJAX UNTUK LOAD PERSONIL BERDASARKAN ASSIGNED_TO (BARU)
# ==============================================================================
def load_personil_by_assigned_to(request):
    """
    AJAX endpoint untuk mengupdate daftar Personil berdasarkan assigned_to user
    """
    try:
        # Cek apakah user sudah login
        if not request.user.is_authenticated:
            return JsonResponse({"error": "User tidak terautentikasi"}, status=401)
        
        assigned_to_id = request.GET.get('assigned_to_id')
        user = request.user
        
        if assigned_to_id and assigned_to_id != '':
            # Jika assigned_to dipilih, ambil personil milik user tersebut
            personil_list = Personil.objects.filter(
                penanggung_jawab_id=assigned_to_id
            ).order_by('nama_lengkap').values('id', 'nama_lengkap')
        else:
            # Jika tidak dipilih, ambil personil milik user yang login
            personil_list = Personil.objects.filter(
                penanggung_jawab=user
            ).order_by('nama_lengkap').values('id', 'nama_lengkap')
        
        data = list(personil_list)
        return JsonResponse(data, safe=False)
    except Exception as e:
        import traceback
        print(f"Error in load_personil_by_assigned_to: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=400)


# ==============================================================================
# VIEW EXPORT DAILY JOBS PDF (BARU)
# ==============================================================================
@login_required(login_url='core:login')
def export_daily_jobs_pdf(request):
    """
    Export Daily Jobs ke PDF dengan filter yang sama seperti di dashboard
    """
    user = request.user
    
    # === 1. LOGIKA FILTER (BULAN & TAHUN) ===
    now = datetime.datetime.now()
    try:
        current_year = int(request.GET.get('year', now.year))
        current_month = int(request.GET.get('month', now.month))
    except ValueError:
        current_year = now.year
        current_month = now.month
    
    # === 2. LOGIKA FILTER PIC ===
    subordinate_ids = user.get_all_subordinates()
    selected_pic_id = request.GET.get('pic', '')
    
    if selected_pic_id == 'my_jobs':
        team_query = Q(pic=user)
    elif selected_pic_id:
        try:
            selected_user = get_object_or_404(CustomUser, id=int(selected_pic_id))
            selected_user_sub_ids = selected_user.get_all_subordinates()
            team_ids = selected_user_sub_ids + [selected_user.id]
            team_query = Q(pic_id__in=team_ids)
        except (ValueError, CustomUser.DoesNotExist):
            team_query = Q(pic=user) | Q(pic_id__in=subordinate_ids)
    else:
        team_query = Q(pic=user) | Q(pic_id__in=subordinate_ids)

    # === 3. LOGIKA FILTER ASET ===
    selected_line_id = request.GET.get('line', '')
    selected_mesin_id = request.GET.get('mesin', '')
    selected_sub_mesin_id = request.GET.get('sub_mesin', '')
    
    # === 4. LOGIKA DATA TABEL ===
    all_jobs_team_base = Job.objects.filter(team_query).distinct()
    
    if selected_sub_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset_id=selected_sub_mesin_id)
    elif selected_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent_id=selected_mesin_id)
    elif selected_line_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent__parent_id=selected_line_id)
    
    # Filter untuk Daily Jobs di bulan/tahun terpilih
    daily_job_data = all_jobs_team_base.filter(
        tipe_job='Daily',
        tanggal_pelaksanaan__tanggal__month=current_month,
        tanggal_pelaksanaan__tanggal__year=current_year
    ).select_related(
        'pic', 
        'project',
        'aset__parent__parent' 
    ).prefetch_related(
        'personil_ditugaskan', 
        'tanggal_pelaksanaan',
        'attachments'
    ).distinct()

    # === 5. HITUNG SUMMARY ===
    summary = calculate_daily_jobs_summary(daily_job_data)
    
    # === 6. TAMBAHKAN ABSOLUTE URL UNTUK ATTACHMENT ===
    request_scheme = request.scheme
    request_host = request.get_host()
    for job in daily_job_data:
        for attachment in job.attachments.all():
            if attachment.file and not attachment.file.url.startswith('http'):
                attachment.absolute_url = f"{request_scheme}://{request_host}{attachment.file.url}"
            else:
                attachment.absolute_url = attachment.file.url if attachment.file else ''
    
    # === 7. SIAPKAN CONTEXT UNTUK TEMPLATE ===
    context = {
        'user': user,
        'job_data': daily_job_data,
        'summary': summary,
        'month_name': get_month_name_id(current_month),
        'year': current_year,
        'report_date': format_tanggal_id(datetime.date.today()),
        'print_date': datetime.datetime.now().strftime("%d %B %Y %H:%M:%S"),
    }
    
    # === 8. RENDER TEMPLATE KE HTML ===
    html_string = render(request, 'report_daily_jobs.html', context).content.decode('utf-8')
    
    # === 9. CONVERT HTML KE PDF DENGAN WEASYPRINT ===
    try:
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        # === 10. GENERATE FILENAME ===
        filename = generate_pdf_filename('daily', current_year, current_month)
        
        # === 11. RETURN PDF FILE ===
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    except Exception as e:
        messages.error(request, f"Gagal generate PDF: {str(e)}")
        return redirect('core:dashboard')


# ==============================================================================
# EXPORT DAILY JOBS TO EXCEL
# ==============================================================================
@login_required(login_url='core:login')
def export_daily_jobs_excel(request):
    user = request.user
    
    # === 1. LOGIKA FILTER (BULAN & TAHUN) ===
    now = datetime.datetime.now()
    try:
        current_year = int(request.GET.get('year', now.year))
        current_month = int(request.GET.get('month', now.month))
    except ValueError:
        current_year = now.year
        current_month = now.month
    
    # === 2. AMBIL DATA JOBS SESUAI FILTER ===
    first_day = datetime.date(current_year, current_month, 1)
    last_day = datetime.date(current_year, current_month, calendar.monthrange(current_year, current_month)[1])
    
    # Ambil job_ids yang punya tanggal dalam range DAN TIDAK PUNYA PROJECT
    job_ids = JobDate.objects.filter(
        tanggal__gte=first_day,
        tanggal__lte=last_day,
        job__project__isnull=True  # Hanya job tanpa project
    ).values_list('job_id', flat=True).distinct()
    
    job_data = Job.objects.filter(id__in=job_ids).order_by('nama_pekerjaan')
    
    # === 3. BUAT WORKBOOK EXCEL ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Jobs"
    
    # === 4. DEFINISIKAN STYLING ===
    header_fill = PatternFill(start_color="2C8C4B", end_color="2C8C4B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === 5. BUAT HEADER ROW ===
    headers = ["No", "Nama Pekerjaan", "PIC", "Line", "Mesin", "Sub", "Fokus", "Prioritas", "Jadwal", "Progress (%)"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # === 6. ISI DATA ===
    for idx, job in enumerate(job_data, 1):
        # Ambil line, mesin, sub dari aset
        line = job.aset.parent.parent.nama if job.aset.parent.parent else "-"
        mesin = job.aset.parent.nama if job.aset.parent else "-"
        sub = job.aset.nama if job.aset else "-"
        
        # Ambil jadwal
        jadwal_dates = job.tanggal_pelaksanaan.filter(
            tanggal__gte=first_day,
            tanggal__lte=last_day
        ).values_list('tanggal', flat=True)
        jadwal_str = ", ".join([str(d.strftime("%d/%m")) for d in jadwal_dates]) if jadwal_dates else "-"
        
        # Ambil progress
        progress = job.get_progress_percent()
        
        row = [
            idx,
            job.nama_pekerjaan,
            job.pic.username if job.pic else "-",
            line,
            mesin,
            sub,
            job.get_fokus_display(),
            job.get_prioritas_display(),
            jadwal_str,
            progress
        ]
        
        ws.append(row)
        
        # Style data row
        for cell in ws[idx + 1]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # === 7. ATUR LEBAR KOLOM ===
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12
    
    # === 8. GENERATE FILENAME ===
    month_name = get_month_name_id(current_month)
    filename = f"Daily_Jobs_{month_name}_{current_year}.xlsx"
    
    # === 9. RETURN EXCEL FILE ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


# ==============================================================================
# EXPORT PROJECT JOBS TO EXCEL
# ==============================================================================
@login_required(login_url='core:login')
def export_project_jobs_excel(request):
    user = request.user
    
    # === 1. LOGIKA FILTER (BULAN & TAHUN) - SAMA SEPERTI DASHBOARD ===
    now = datetime.datetime.now()
    try:
        current_year_param = request.GET.get('year', '')
        current_month_param = request.GET.get('month', '')
        
        # Jika kosong atau '0', gunakan nilai default
        if current_year_param and current_year_param != '0':
            current_year = int(current_year_param)
        else:
            current_year = 0  # Artinya "Semua tahun"
        
        if current_month_param and current_month_param != '0':
            current_month = int(current_month_param)
            # Jika user pilih bulan tapi tidak pilih tahun, default ke tahun sekarang
            if current_year == 0:
                current_year = now.year
        else:
            current_month = 0  # Artinya "Semua bulan"
    except ValueError:
        current_year = 0
        current_month = 0
    
    # Flag untuk menentukan apakah filter aktif atau "Semua"
    filter_all_dates = (current_month == 0 and current_year == 0)
    
    # === 2. LOGIKA FILTER PIC ===
    subordinate_ids = user.get_all_subordinates()
    selected_pic_id = request.GET.get('pic', '')
    
    if selected_pic_id == 'my_jobs':
        team_query = Q(pic=user)
    elif selected_pic_id:
        try:
            selected_user = get_object_or_404(CustomUser, id=int(selected_pic_id))
            selected_user_sub_ids = selected_user.get_all_subordinates()
            team_ids = selected_user_sub_ids + [selected_user.id]
            team_query = Q(pic_id__in=team_ids)
        except (ValueError, CustomUser.DoesNotExist):
            team_query = Q(pic=user) | Q(pic_id__in=subordinate_ids)
    else:
        team_query = Q(pic=user) | Q(pic_id__in=subordinate_ids)
    
    # === 3. LOGIKA FILTER ASET ===
    selected_line_id = request.GET.get('line', '')
    selected_mesin_id = request.GET.get('mesin', '')
    selected_sub_mesin_id = request.GET.get('sub_mesin', '')
    
    # === 4. LOGIKA DATA TABEL ===
    all_jobs_team_base = Job.objects.filter(team_query, tipe_job='Project').distinct()
    
    if selected_sub_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset_id=selected_sub_mesin_id)
    elif selected_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent_id=selected_mesin_id)
    elif selected_line_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent__parent_id=selected_line_id)
    
    # Filter berdasarkan bulan dan tahun (sama seperti dashboard)
    if filter_all_dates:
        # Ambil semua project jobs tanpa filter tanggal
        project_jobs = all_jobs_team_base.select_related(
            'pic', 
            'project',
            'aset',
            'aset__parent',
            'aset__parent__parent'
        ).prefetch_related('tanggal_pelaksanaan').order_by('project__nama_project', 'nama_pekerjaan').distinct()
    else:
        # Build filter dynamically based on what user selected
        date_filter = Q()
        if current_month != 0:
            date_filter &= Q(tanggal_pelaksanaan__tanggal__month=current_month)
        if current_year != 0:
            date_filter &= Q(tanggal_pelaksanaan__tanggal__year=current_year)
        
        project_jobs = all_jobs_team_base.filter(date_filter).select_related(
            'pic', 
            'project',
            'aset',
            'aset__parent',
            'aset__parent__parent'
        ).prefetch_related('tanggal_pelaksanaan').order_by('project__nama_project', 'nama_pekerjaan').distinct()
    
    # === 5. BUAT WORKBOOK EXCEL ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Jobs"
    
    # === 6. DEFINISIKAN STYLING ===
    header_fill = PatternFill(start_color="2C8C4B", end_color="2C8C4B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    project_fill = PatternFill(start_color="D4E6DC", end_color="D4E6DC", fill_type="solid")
    project_font = Font(bold=True, size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === 7. BUAT HEADER ROW ===
    headers = ["No", "Project", "Nama Pekerjaan", "PIC", "Line", "Mesin", "Sub", "Fokus", "Prioritas", "Jadwal", "Progress (%)"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # === 8. ISI DATA PER PROJECT ===
    row_num = 1
    no = 0
    
    # Group jobs by project
    project_dict = {}
    for job in project_jobs:
        project_name = job.project.nama_project if job.project else "Tanpa Project"
        if project_name not in project_dict:
            project_dict[project_name] = {
                'project': job.project,
                'jobs': []
            }
        project_dict[project_name]['jobs'].append(job)
    
    for project_name, data in project_dict.items():
        project = data['project']
        jobs = data['jobs']
        
        # Buat project header
        if project:
            row_num += 1
            ws.insert_rows(row_num)
            ws[row_num][0].value = ""
            ws[row_num][1].value = f"PROJECT: {project.nama_project}"
            
            # Merge cells untuk project name
            ws.merge_cells(f'B{row_num}:J{row_num}')
            
            # Style project header
            for col in range(1, 11):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = project_fill
                cell.font = project_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Isi job data untuk project ini
        for job in jobs:
            no += 1
            row_num += 1
            
            # Ambil line, mesin, sub dari aset
            line = job.aset.parent.parent.nama if job.aset and job.aset.parent and job.aset.parent.parent else "-"
            mesin = job.aset.parent.nama if job.aset and job.aset.parent else "-"
            sub = job.aset.nama if job.aset else "-"
            
            # Ambil jadwal (dengan filter bulan/tahun jika ada)
            if filter_all_dates:
                # Tampilkan semua jadwal
                jadwal_dates = job.tanggal_pelaksanaan.all().values_list('tanggal', flat=True).order_by('tanggal')
            else:
                # Ambil jadwal sesuai filter bulan/tahun
                date_filter_jadwal = Q()
                if current_month != 0:
                    date_filter_jadwal &= Q(tanggal__month=current_month)
                if current_year != 0:
                    date_filter_jadwal &= Q(tanggal__year=current_year)
                jadwal_dates = job.tanggal_pelaksanaan.filter(date_filter_jadwal).values_list('tanggal', flat=True).order_by('tanggal')
            
            jadwal_str = ", ".join([str(d.strftime("%d/%m")) for d in jadwal_dates]) if jadwal_dates else "-"
            
            # Ambil progress
            progress = job.get_progress_percent()
            
            row = [
                no,
                "",
                job.nama_pekerjaan,
                job.pic.username if job.pic else "-",
                line,
                mesin,
                sub,
                job.get_fokus_display(),
                job.get_prioritas_display(),
                jadwal_str,
                progress
            ]
            
            ws.append(row)
            
            # Style data row
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # === 9. ATUR LEBAR KOLOM ===
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 12
    
    # === 8. GENERATE FILENAME ===
    month_name = get_month_name_id(current_month)
    filename = f"Project_Jobs_{month_name}_{current_year}.xlsx"
    
    # === 9. RETURN EXCEL FILE ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

# ==============================================================================
# VIEW EXPORT PROJECT JOBS PDF (BARU)
# ==============================================================================
@login_required(login_url='core:login')
def export_project_jobs_pdf(request):
    """
    Export Project Jobs ke PDF dengan filter yang sama seperti di dashboard
    """
    user = request.user
    
    # === 1. LOGIKA FILTER (BULAN & TAHUN) - SAMA SEPERTI DASHBOARD ===
    now = datetime.datetime.now()
    try:
        current_year_param = request.GET.get('year', '')
        current_month_param = request.GET.get('month', '')
        
        # Jika kosong atau '0', gunakan nilai default
        if current_year_param and current_year_param != '0':
            current_year = int(current_year_param)
        else:
            current_year = 0  # Artinya "Semua tahun"
        
        if current_month_param and current_month_param != '0':
            current_month = int(current_month_param)
            # Jika user pilih bulan tapi tidak pilih tahun, default ke tahun sekarang
            if current_year == 0:
                current_year = now.year
        else:
            current_month = 0  # Artinya "Semua bulan"
    except ValueError:
        current_year = 0
        current_month = 0
    
    # Flag untuk menentukan apakah filter aktif atau "Semua"
    filter_all_dates = (current_month == 0 and current_year == 0)
    
    # === 2. LOGIKA FILTER PIC ===
    subordinate_ids = user.get_all_subordinates()
    selected_pic_id = request.GET.get('pic', '')
    
    if selected_pic_id == 'my_jobs':
        team_query = Q(pic=user)
    elif selected_pic_id:
        try:
            selected_user = get_object_or_404(CustomUser, id=int(selected_pic_id))
            selected_user_sub_ids = selected_user.get_all_subordinates()
            team_ids = selected_user_sub_ids + [selected_user.id]
            team_query = Q(pic_id__in=team_ids)
        except (ValueError, CustomUser.DoesNotExist):
            team_query = Q(pic=user) | Q(pic_id__in=subordinate_ids)
    else:
        team_query = Q(pic=user) | Q(pic_id__in=subordinate_ids)

    # === 3. LOGIKA FILTER ASET ===
    selected_line_id = request.GET.get('line', '')
    selected_mesin_id = request.GET.get('mesin', '')
    selected_sub_mesin_id = request.GET.get('sub_mesin', '')
    
    # === 4. LOGIKA DATA TABEL ===
    all_jobs_team_base = Job.objects.filter(team_query).distinct()
    
    if selected_sub_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset_id=selected_sub_mesin_id)
    elif selected_mesin_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent_id=selected_mesin_id)
    elif selected_line_id:
        all_jobs_team_base = all_jobs_team_base.filter(aset__parent__parent_id=selected_line_id)
    
    # Filter untuk Project Jobs (sama seperti dashboard)
    if filter_all_dates:
        # Ambil semua project jobs tanpa filter tanggal
        project_jobs = all_jobs_team_base.filter(
            tipe_job='Project'
        ).select_related(
            'pic', 
            'project',
            'aset__parent__parent' 
        ).prefetch_related(
            'personil_ditugaskan', 
            'tanggal_pelaksanaan',
            'attachments'
        ).order_by('project__nama_project', 'nama_pekerjaan').distinct()
    else:
        # Build filter dynamically based on what user selected
        date_filter = Q()
        if current_month != 0:
            date_filter &= Q(tanggal_pelaksanaan__tanggal__month=current_month)
        if current_year != 0:
            date_filter &= Q(tanggal_pelaksanaan__tanggal__year=current_year)
        
        project_jobs = all_jobs_team_base.filter(
            tipe_job='Project'
        ).filter(date_filter).select_related(
            'pic', 
            'project',
            'aset__parent__parent' 
        ).prefetch_related(
            'personil_ditugaskan', 
            'tanggal_pelaksanaan',
            'attachments'
        ).order_by('project__nama_project', 'nama_pekerjaan').distinct()

    # === 5. GRUP JOBS BERDASARKAN PROJECT ===
    project_dict = {}
    for job in project_jobs:
        if job.project:
            if job.project.id not in project_dict:
                project_dict[job.project.id] = {
                    'project': job.project,
                    'jobs': []
                }
            project_dict[job.project.id]['jobs'].append(job)
    
    # === 6. HITUNG PROGRESS PER PROJECT ===
    project_data = []
    for project_id, data in project_dict.items():
        jobs = data['jobs']
        total_dates = sum(job.tanggal_pelaksanaan.count() for job in jobs)
        done_dates = sum(job.tanggal_pelaksanaan.filter(status='Done').count() for job in jobs)
        
        progress = 0
        if total_dates > 0:
            progress = int((done_dates / total_dates) * 100)
        
        project_data.append({
            'project': data['project'],
            'jobs': jobs,
            'progress': progress
        })
    
    # === 7. HITUNG SUMMARY ===
    summary = calculate_project_jobs_summary(project_data)
    
    # === 8. TAMBAHKAN ABSOLUTE URL UNTUK ATTACHMENT ===
    request_scheme = request.scheme
    request_host = request.get_host()
    for project_item in project_data:
        for job in project_item['jobs']:
            for attachment in job.attachments.all():
                if attachment.file and not attachment.file.url.startswith('http'):
                    attachment.absolute_url = f"{request_scheme}://{request_host}{attachment.file.url}"
                else:
                    attachment.absolute_url = attachment.file.url if attachment.file else ''
    
    # === 9. SIAPKAN CONTEXT UNTUK TEMPLATE ===
    context = {
        'user': user,
        'project_data': project_data,
        'summary': summary,
        'month_name': get_month_name_id(current_month),
        'year': current_year,
        'report_date': format_tanggal_id(datetime.date.today()),
        'print_date': datetime.datetime.now().strftime("%d %B %Y %H:%M:%S"),
    }
    
    # === 10. RENDER TEMPLATE KE HTML ===
    html_string = render(request, 'report_project_jobs.html', context).content.decode('utf-8')
    
    # === 11. CONVERT HTML KE PDF DENGAN WEASYPRINT ===
    try:
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        # === 12. GENERATE FILENAME ===
        filename = generate_pdf_filename('project', current_year, current_month)
        
        # === 13. RETURN PDF FILE ===
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    except Exception as e:
        messages.error(request, f"Gagal generate PDF: {str(e)}")
        return redirect('core:dashboard')


# ==============================================================================
# API EXPORT - SEND DATA TO GOOGLE APPS SCRIPT
# ==============================================================================
@login_required(login_url='core:login')
def export_jobs_to_gas(request):
    """
    API endpoint untuk export selected jobs ke Google Apps Script
    Expected POST data:
    {
        "job_ids": [1, 2, 3, ...],
        "export_type": "preventif" atau "evaluasi"
    }
    """
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Method not allowed"}, status=405)
    
    try:
        # Parse request data
        data = json.loads(request.body)
        job_ids = data.get('job_ids', [])
        export_type = data.get('export_type', 'preventif')
        
        if not job_ids:
            return JsonResponse({"status": "error", "message": "Tidak ada job yang dipilih"})
        
        # Import handler functions
        from .export_handlers import prepare_job_data_for_export, send_to_google_apps_script
        
        # Prepare data
        payload = prepare_job_data_for_export(job_ids)
        if not payload:
            return JsonResponse({"status": "error", "message": "Job tidak ditemukan"})
        
        # Send to Google Apps Script
        result = send_to_google_apps_script(payload, export_type)
        
        return JsonResponse(result)
    
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "Invalid JSON"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Error: {str(e)}"})


# ==============================================================================
# VIEW JOB PER DAY - UNTUK EXPORT KE GOOGLE APPS SCRIPT
# ==============================================================================
@login_required(login_url='core:login')
def job_per_day_view(request):
    """
    Halaman untuk melihat jobs per hari dan export ke Google Apps Script (Preventif/Evaluasi)
    User bisa select multiple jobs dalam satu hari, kemudian export
    Format: TABEL dengan checkbox, filter (Line, Mesin, Sub Mesin)
    """
    import datetime
    from django.db.models import Q
    
    user = request.user
    now = datetime.datetime.now()
    
    # Get filter parameters
    selected_date_str = request.GET.get('date', now.strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = now.date()
    
    # Get user's subordinates untuk permission check
    subordinate_ids = user.get_all_subordinates()
    
    # Filter jobs untuk tanggal tertentu
    # User bisa lihat job jika:
    # 1. User adalah PIC
    # 2. User adalah assigned_to
    # 3. User adalah supervisor dari PIC atau assigned_to
    jobs_for_date = JobDate.objects.filter(
        tanggal=selected_date
    ).select_related('job').prefetch_related(
        'job__personil_ditugaskan',
        'job__aset',
        'job__aset__parent',
        'job__aset__parent__parent'
    )
    
    # Filter berdasarkan permission
    visible_jobs = []
    for job_date in jobs_for_date:
        job = job_date.job
        can_view = (
            job.pic == user or
            job.assigned_to == user or
            (job.pic and job.pic.id in subordinate_ids) or
            (job.assigned_to is not None and job.assigned_to.id in subordinate_ids)
        )
        if can_view:
            visible_jobs.append(job)
    
    # Remove duplicates (dalam case ada multiple JobDate untuk job yang sama)
    visible_jobs = list(set(visible_jobs))
    
    # Collect all unique Line, Mesin, Sub Mesin untuk dropdown filters
    all_lines = set()
    all_mesins = set()
    all_submesins = set()
    
    for job in visible_jobs:
        if job.aset:
            if job.aset.parent and job.aset.parent.parent:
                all_lines.add(job.aset.parent.parent.nama)
            if job.aset.parent:
                all_mesins.add(job.aset.parent.nama)
            all_submesins.add(job.aset.nama)
    
    # Sort untuk dropdown
    all_lines = sorted(list(all_lines))
    all_mesins = sorted(list(all_mesins))
    all_submesins = sorted(list(all_submesins))
    
    context = {
        'selected_date': selected_date,
        'all_jobs': visible_jobs,
        'all_lines': all_lines,
        'all_mesins': all_mesins,
        'all_submesins': all_submesins,
        'total_jobs': len(visible_jobs),
    }
    
    return render(request, 'job_per_day.html', context)


# ==============================================================================
# API UNTUK FETCH ATTACHMENTS GALLERY (BARU)
# ==============================================================================
@login_required(login_url='core:login')
def api_job_attachments(request, job_id):
    """
    API endpoint untuk fetch semua attachments dari sebuah job.
    Return format JSON dengan list attachment objects.
    """
    try:
        user = request.user
        job = get_object_or_404(Job, id=job_id)
        
        # === PERMISSION CHECK ===
        subordinate_ids = user.get_all_subordinates()
        can_view = (
            job.pic == user or
            job.assigned_to == user or
            (job.pic and job.pic.id in subordinate_ids) or
            (job.assigned_to is not None and job.assigned_to.id in subordinate_ids)
        )
        
        if not can_view:
            return JsonResponse({"error": "Akses ditolak"}, status=403)
        
        # Fetch semua attachments
        attachments = job.attachments.all()
        
        # Convert ke JSON
        attachments_data = []
        for att in attachments:
            attachments_data.append({
                'id': att.id,
                'file_name': att.file.name.split('/')[-1] if att.file else 'Unnamed',
                'file_url': att.file.url if att.file else '',
                'deskripsi': att.deskripsi or '',
                'tipe_file': att.tipe_file,
            })
        
        return JsonResponse({
            'job_id': job_id,
            'job_name': job.nama_pekerjaan,
            'attachments': attachments_data,
        })
    
    except Exception as e:
        import traceback
        print(f"Error in api_job_attachments: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=400)


# ==============================================================================
# VIEW LEAVE EVENT (IJIN/CUTI) - HALAMAN BARU
# ==============================================================================
def sync_leave_events_from_google_calendar():
    """
    Sync events dari Google Calendar ke database
    Baca semua events yang nama-nya mengandung "cuti" dan simpan ke DB jika belum ada
    """
    try:
        from .google_calendar_service import get_google_calendar_service
        from datetime import datetime
        
        cal_service = get_google_calendar_service()
        calendar_id = settings.GOOGLE_CALENDAR_ID
        
        # Query semua events dengan keyword "cuti"
        events_result = cal_service.service.events().list(
            calendarId=calendar_id,
            q='cuti',  # Search untuk events dengan "cuti" di nama
            maxResults=100
        ).execute()
        
        events = events_result.get('items', [])
        synced_count = 0
        
        for event in events:
            event_id = event.get('id')
            summary = event.get('summary', '')
            
            # Skip jika sudah ada di database
            if LeaveEvent.objects.filter(google_event_id=event_id).exists():
                continue
            
            # Parse data dari event
            # Format: "{Nama} cuti"
            nama_orang = summary.replace(' cuti', '').strip()
            
            # Get tanggal dari event
            start = event.get('start', {})
            end = event.get('end', {})
            
            start_date = start.get('date') or start.get('dateTime', '')
            end_date = end.get('date') or end.get('dateTime', '')
            
            if start_date and end_date:
                # Simplify: hanya ambil date part
                start_date = start_date[:10]
                end_date = end_date[:10]
                
                # Cari karyawan dengan nama yang match
                try:
                    karyawan = Karyawan.objects.get(nama_lengkap__icontains=nama_orang)
                except Karyawan.DoesNotExist:
                    karyawan = None
                
                # Create atau update LeaveEvent di database
                leave_event, created = LeaveEvent.objects.get_or_create(
                    google_event_id=event_id,
                    defaults={
                        'karyawan': karyawan,
                        'nama_orang': nama_orang,
                        'tipe_leave': 'Cuti',
                        'tanggal': start_date,  # Atau bisa range jika perlu
                        'deskripsi': event.get('description', ''),
                        'created_by': None,
                    }
                )
                
                if created:
                    synced_count += 1
        
        return synced_count
        
    except Exception as e:
        print(f"Error syncing from Google Calendar: {e}")
        return 0


@login_required(login_url='core:login')
def leave_event_view(request):
    """
    Halaman untuk create/manage Leave Events (Ijin/Cuti)
    dengan tab untuk cuti yang akan datang dan yang sudah kelewat
    """
    user = request.user
    
    # Sync data dari Google Calendar
    sync_leave_events_from_google_calendar()
    
    if request.method == 'POST':
        form = LeaveEventForm(request.POST)
        
        if form.is_valid():
            try:
                # Get form data
                karyawan = form.cleaned_data['karyawan']
                nama_orang = karyawan.nama_lengkap
                tipe_leave = form.cleaned_data['tipe_leave']
                tanggal_str = form.cleaned_data['tanggal_picker']
                deskripsi = form.cleaned_data.get('deskripsi', '')
                
                # Parse tanggal string menjadi list
                tanggal_list = [tgl.strip() for tgl in tanggal_str.split(',') if tgl.strip()]
                
                # === PUSH KE GOOGLE CALENDAR ===
                from .google_calendar_service import get_google_calendar_service
                
                try:
                    cal_service = get_google_calendar_service()
                    event_response = cal_service.create_event(
                        nama_orang=nama_orang,
                        tipe_ijin=tipe_leave,
                        tanggal_list=tanggal_list,
                        deskripsi=deskripsi
                    )
                    
                    if event_response:
                        event_id = event_response.get('id')
                        
                        # === SIMPAN KE DATABASE ===
                        leave_event = form.save(commit=False)
                        leave_event.google_event_id = event_id
                        leave_event.created_by = user
                        leave_event.save()
                        
                        messages.success(
                            request, 
                            f"✅ Sukses! Event '{nama_orang} cuti' telah ditambahkan ke Google Calendar."
                        )
                        return redirect('core:leave_event')
                    else:
                        messages.error(
                            request,
                            "❌ Gagal menambahkan event ke Google Calendar. Cek error di server logs."
                        )
                
                except Exception as e:
                    messages.error(
                        request,
                        f"❌ Error koneksi Google Calendar: {str(e)}"
                    )
                    print(f"Google Calendar Error: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
            except Exception as e:
                messages.error(request, f"❌ Error: {str(e)}")
                import traceback
                traceback.print_exc()
        else:
            # Form tidak valid
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    
    else:
        form = LeaveEventForm()
    
    # Ambil daftar leave events yang pernah dibuat
    all_leave_events = LeaveEvent.objects.all().order_by('-created_at')
    
    # Split data ke upcoming vs past (berdasarkan tanggal terakhir)
    from datetime import datetime as dt
    today = dt.now().date()
    
    upcoming_events = []
    past_events = []
    
    for event in all_leave_events:
        tanggal_list = event.get_tanggal_list()
        if tanggal_list:
            last_date = sorted(tanggal_list)[-1]
            try:
                last_date_obj = dt.strptime(last_date, '%Y-%m-%d').date()
                if last_date_obj >= today:
                    upcoming_events.append(event)
                else:
                    past_events.append(event)
            except:
                upcoming_events.append(event)
        else:
            upcoming_events.append(event)
    
    # Build calendar data untuk template
    import calendar as cal_module
    from datetime import datetime, timedelta
    
    now = datetime.now()
    
    # Get month/year dari URL params jika ada
    try:
        current_month = int(request.GET.get('month', now.month))
        current_year = int(request.GET.get('year', now.year))
    except (ValueError, TypeError):
        current_month = now.month
        current_year = now.year
    
    # Validate month/year
    if current_month < 1:
        current_month = 12
        current_year -= 1
    elif current_month > 12:
        current_month = 1
        current_year += 1
    
    # Get hari pertama bulan dan jumlah hari
    first_day_weekday, num_days = cal_module.monthrange(current_year, current_month)
    
    # Build list of dates dengan leave info
    calendar_data = []
    
    # Padding hari kosong awal bulan
    for _ in range(first_day_weekday):
        calendar_data.append({'date': None, 'leave_events': []})
    
    # Fill tanggal dengan data
    for day in range(1, num_days + 1):
        date_str = f"{current_year}-{current_month:02d}-{day:02d}"
        
        # Cari leave events yang match tanggal ini
        day_leaves = []
        for event in all_leave_events:
            tanggal_list = event.get_tanggal_list()
            if date_str in tanggal_list:
                day_leaves.append(event)
        
        calendar_data.append({
            'date': day,
            'date_str': date_str,
            'leave_events': day_leaves
        })
    
    # Get month name
    month_names = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                   'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    month_name = month_names[current_month - 1]
    
    # Get karyawan list untuk datalist
    karyawan_list = Karyawan.objects.filter(status='Aktif').order_by('nama_lengkap')
    
    context = {
        'form': form,
        'karyawan_list': karyawan_list,
        'all_leave_events': all_leave_events,
        'upcoming_events': upcoming_events,
        'past_events': past_events,
        'calendar_data': calendar_data,
        'calendar_rows': [calendar_data[i:i+7] for i in range(0, len(calendar_data), 7)],
        'current_month': current_month,
        'current_year': current_year,
        'month_name': month_name,
    }
    
    return render(request, 'leave_event.html', context)


@login_required(login_url='core:login')
def leave_event_list(request):
    """
    Halaman daftar Leave Events
    """
    leave_events = LeaveEvent.objects.all().order_by('-created_at')
    
    context = {
        'leave_events': leave_events,
    }
    
    return render(request, 'leave_event_list.html', context)


@login_required(login_url='core:login')
def leave_event_detail(request, leave_id):
    """
    Halaman detail Leave Event
    """
    leave_event = get_object_or_404(LeaveEvent, id=leave_id)
    
    context = {
        'leave_event': leave_event,
    }
    
    return render(request, 'leave_event_detail.html', context)


@login_required(login_url='core:login')
def leave_event_delete(request, leave_id):
    """
    Delete Leave Event (juga delete dari Google Calendar)
    """
    leave_event = get_object_or_404(LeaveEvent, id=leave_id)
    
    try:
        # Delete dari Google Calendar jika ada event_id
        if leave_event.google_event_id:
            from .google_calendar_service import get_google_calendar_service
            cal_service = get_google_calendar_service()
            cal_service.delete_event(leave_event.google_event_id)
        
        # Delete dari database
        nama = leave_event.nama_orang
        leave_event.delete()
        
        messages.success(request, f"✅ Sukses menghapus event '{nama}' dari database dan Google Calendar.")
        
    except Exception as e:
        messages.error(request, f"❌ Error saat delete: {str(e)}")
        print(f"Delete Error: {str(e)}")
    
    return redirect('core:leave_event_list')


# ==============================================================================
# TEMPORARY: VIEW UNTUK RUN MIGRATION (HELPER SAJA)
# ==============================================================================
def run_migration_helper(request):
    """
    Temporary view untuk apply pending migrations
    HANYA untuk development/setup. HAPUS di production!
    """
    from django.core.management import call_command
    from io import StringIO
    
    # Security: hanya jalan jika DEBUG=True atau dari localhost
    if not (settings.DEBUG or request.META.get('REMOTE_ADDR') in ['127.0.0.1', 'localhost']):
        return HttpResponse("Forbidden", status=403)
    
    try:
        out = StringIO()
        call_command('migrate', stdout=out)
        
        result = f"""
        <h2>✅ Migration Success</h2>
        <pre>{out.getvalue()}</pre>
        <a href="/">Back to Dashboard</a>
        """
        return HttpResponse(result)
    except Exception as e:
        return HttpResponse(f"<h2>❌ Error: {str(e)}</h2><pre>{str(e)}</pre>", status=500)

